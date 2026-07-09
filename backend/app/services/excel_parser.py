from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl

REQUIRED_COLUMNS = ["item_type", "japanese", "kana", "romaji", "meanings", "part_of_speech"]
ITEMS_SHEET_NAME = "items"
VALID_ITEM_TYPES = ("word", "phrase")


class WordBankFileError(Exception):
    """A file-level problem that blocks the entire file from importing."""


@dataclass
class ParsedExample:
    japanese: str
    kana: str
    english: str


@dataclass
class ParsedRow:
    row_number: int
    item_type: str
    japanese: str
    kana: str
    romaji: str
    meanings: List[str]
    part_of_speech: str
    examples: List[ParsedExample]
    similar_items: List[str]
    source_note: Optional[str]


@dataclass
class RowError:
    row_number: int
    message: str


@dataclass
class ParsedWorkbook:
    rows: List[ParsedRow]
    errors: List[RowError]


def _split_semicolon(value) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_workbook(path: Path) -> ParsedWorkbook:
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any load failure blocks the whole file
        raise WordBankFileError(f"Could not open '{path.name}' as an .xlsx file: {exc}") from exc

    if ITEMS_SHEET_NAME not in workbook.sheetnames:
        raise WordBankFileError(f"'{path.name}' is missing the required '{ITEMS_SHEET_NAME}' sheet.")

    worksheet = workbook[ITEMS_SHEET_NAME]
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise WordBankFileError(f"'{path.name}' items sheet has no header row.") from exc

    header = [_cell_str(cell) for cell in header_row]
    col_index = {name: idx for idx, name in enumerate(header) if name}

    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        raise WordBankFileError(
            f"'{path.name}' items sheet is missing required column(s): {', '.join(missing)}."
        )

    rows: List[ParsedRow] = []
    errors: List[RowError] = []

    for offset, raw_row in enumerate(rows_iter, start=2):  # header occupies row 1
        if raw_row is None or all(cell is None for cell in raw_row):
            continue

        def get(col: str):
            idx = col_index.get(col)
            if idx is None or idx >= len(raw_row):
                return None
            return raw_row[idx]

        item_type = _cell_str(get("item_type"))
        japanese = _cell_str(get("japanese"))
        kana = _cell_str(get("kana"))
        romaji = _cell_str(get("romaji"))
        part_of_speech = _cell_str(get("part_of_speech"))
        meanings = _split_semicolon(get("meanings"))

        row_errors = []
        if item_type not in VALID_ITEM_TYPES:
            row_errors.append("item_type must be 'word' or 'phrase'")
        if not japanese:
            row_errors.append("japanese is required")
        if not kana:
            row_errors.append("kana is required")
        if not romaji:
            row_errors.append("romaji is required")
        if not part_of_speech:
            row_errors.append("part_of_speech is required")
        if not meanings:
            row_errors.append("meanings must contain at least one value")

        example_japanese = _split_semicolon(get("example_japanese"))
        example_kana = _split_semicolon(get("example_kana"))
        example_english = _split_semicolon(get("example_english"))
        if len({len(example_japanese), len(example_kana), len(example_english)}) > 1:
            row_errors.append(
                "example_japanese, example_kana, and example_english must have matching counts"
            )

        if row_errors:
            errors.append(RowError(row_number=offset, message="; ".join(row_errors)))
            continue

        examples = [
            ParsedExample(japanese=ja, kana=ka, english=en)
            for ja, ka, en in zip(example_japanese, example_kana, example_english)
        ]

        similar_items = _split_semicolon(get("similar_items"))
        source_note_raw = get("source_note")
        source_note = _cell_str(source_note_raw) or None

        rows.append(
            ParsedRow(
                row_number=offset,
                item_type=item_type,
                japanese=japanese,
                kana=kana,
                romaji=romaji,
                meanings=meanings,
                part_of_speech=part_of_speech,
                examples=examples,
                similar_items=similar_items,
                source_note=source_note,
            )
        )

    return ParsedWorkbook(rows=rows, errors=errors)
